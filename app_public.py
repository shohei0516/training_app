from pathlib import Path
from html import escape
import base64
import mimetypes
import re

import qrcode
import streamlit as st
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "pitching_form_training.xlsx"

ASSETS_DIR = BASE_DIR / "assets"
GIF_DIR = ASSETS_DIR / "eval_gif"
OK_NG_DIR = ASSETS_DIR / "ok_ng"
PHASE_IMAGE_DIR = ASSETS_DIR / "phase_images"

TRAINING_GIF_DIRS = [
    BASE_DIR / "gif",
    ASSETS_DIR / "gif",
    ASSETS_DIR / "training_gif",
    BASE_DIR / "training_video" / "GIF",
]

OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

PHASE_IMAGE_MAP = {
    "ワインドアップ": "windup.png",
    "アーリーコッキング": "early_cocking.png",
    "レイトコッキング": "late_cocking.png",
    "アクセラレーション": "acceleration.png",
    "フォロースルー": "follow_through.png",
}


CHECKS = [
    {"id": "C01", "phase": "ワインドアップ", "title": "軸足安定", "criteria": {0: "・骨盤水平・体幹ブレなし\n・支持脚で安定して立位保持できる", 1: "・骨盤がわずかに傾く\n・または体幹の軽度左右ブレあり", 2: "・骨盤の横シフトが大きい\n・または明らかな体幹動揺がある"}},
    {"id": "C02", "phase": "アーリーコッキング", "title": "軸足膝前方移動", "criteria": {0: "・軸足膝がつま先と同程度\n・またはやや後方に位置している", 1: "・軸足膝がつま先よりわずかに前方へ出る", 2: "・軸足膝がつま先より明らかに前方へ突出する\n・前方移動優位になっている"}},
    {"id": "C03", "phase": "アーリーコッキング", "title": "軸足アライメント", "criteria": {0: "・股関節・膝・足部のラインが一直線\n・Knee inがみられない", 1: "・膝がわずかに内側へ入る\n・ただしラインの崩れは軽度", 2: "・膝が明らかに内側へ崩れる\n・股関節・膝・足部の一直線が保てない"}},
    {"id": "C04", "phase": "レイトコッキング", "title": "ステップ足膝屈曲", "criteria": {0: "・接地時の膝屈曲が適度\n・約40〜60°程度", 1: "・接地時の膝屈曲がやや大きい\n・沈み込みがやや目立つ", 2: "・接地時に膝が深く曲がりすぎる\n・約100°以上の過度屈曲"}},
    {"id": "C05", "phase": "レイトコッキング", "title": "ステップ足膝外側動揺", "criteria": {0: "・ステップ足が進行方向に対して直線的に接地している", 1: "・ステップ足がわずかに外側へ開く\n・または回転時に軽度の外逃げがある", 2: "・ステップ足が明らかに外側へ逃げる\n・回転運動中に支持が不安定になる"}},
    {"id": "C06", "phase": "レイトコッキング", "title": "テイクバック位置", "criteria": {0: "・テイクバックでボールが頭部後方に位置する\n・正面から見て頭の後ろに隠れている", 1: "・ボールがやや横に見える\n・頭部後方への引き込みがやや不足する", 2: "・ボールが横または前方に明らかに見える"}},
    {"id": "C07", "phase": "アクセラレーション", "title": "Cカーブ形成", "criteria": {0: "・肩最大外旋位で体全体がCカーブを描く", 1: "・Cカーブがやや不十分", 2: "・Cカーブが明らかに形成できていない\n・体のラインが一直線または前屈傾向"}},
    {"id": "C08", "phase": "アクセラレーション", "title": "ステップ足膝伸展", "criteria": {0: "・ステップ足接地後からリリースまで膝が伸展する", 1: "・膝伸展がやや不十分\n・接地後の膝伸び上がりが弱い", 2: "・膝屈曲位を維持したままリリースする"}},
    {"id": "C09", "phase": "アクセラレーション", "title": "肘抜け", "criteria": {0: "・頭部の横でリリースできる\n・肩・肘・手のラインが一直線に近い", 1: "・肘がやや先行する\n・リリース位置がわずかに前方へずれる", 2: "・肘が明らかに先行する\n・頭部より前方でリリースし肘抜けがみられる"}},
    {"id": "C10", "phase": "フォロースルー", "title": "上体回旋角度", "criteria": {0: "・両肩を結んだ線がセンターとキャッチャー方向を向く", 1: "・回旋はあるがやや不十分\n・肩のラインが目標方向まで回りきらない", 2: "・回旋が途中で止まる\n・肩のラインが明らかに目標方向を向かない"}},
    {"id": "C11", "phase": "フォロースルー", "title": "軸足けり上げ", "criteria": {0: "・軸足が後方へ振り上がる\n・足部が腰のラインよりも高い", 1: "・軸足が振り上がるが腰のラインよりも低い", 2: "・軸足が残る、もしくは引きずる\n・蹴り上げがほとんどみられない"}},
]


def image_to_base64_src(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    mime_type, _ = mimetypes.guess_type(path)
    if mime_type is None:
        mime_type = "image/png"
    data = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:{mime_type};base64,{data}"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_ids(value: str) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [x.strip() for x in re.split(r"[,、/\s\n]+", text) if x.strip()]


def get_cell(row_dict: dict, candidates: list[str]) -> str:
    for key, value in row_dict.items():
        key_text = normalize_text(key)
        for cand in candidates:
            if cand in key_text:
                return normalize_text(value)
    return ""


def find_training_gif(training_id: str, filename: str = "") -> Path | None:
    names = []

    if filename:
        names.append(filename)
        if not filename.lower().endswith(".gif"):
            names.append(filename + ".gif")

    names.append(f"{training_id}.gif")

    for folder in TRAINING_GIF_DIRS:
        for name in names:
            p = folder / name
            if p.exists():
                return p

    return None


def read_sheet_as_dicts(wb, sheet_name: str) -> list[dict]:
    ws = wb[sheet_name]
    headers = [normalize_text(c.value) for c in ws[1]]
    rows = []

    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for i, header in enumerate(headers, start=1):
            value = ws.cell(r, i).value
            if value not in [None, ""]:
                empty = False
            row[header] = value
        if not empty:
            rows.append(row)

    return rows


def load_training_data() -> tuple[dict, dict]:
    """
    Excelから以下を読む。
    ① トレーニングマスタ
    ② チェックIDとトレーニングIDの紐づけ表

    シート名や列名が多少違っても拾えるようにしている。
    """

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, data_only=True)

    training_sheet = None
    mapping_sheet = None

    for s in wb.sheetnames:
        if "トレーニングマスタ" in s or "training" in s.lower():
            training_sheet = s
        if "紐" in s or "対応" in s or "リンク" in s or "連動" in s:
            mapping_sheet = s

    if training_sheet is None:
        for s in wb.sheetnames:
            rows = read_sheet_as_dicts(wb, s)
            if not rows:
                continue
            headers = rows[0].keys()
            header_text = " ".join(headers)
            if "トレーニング" in header_text and ("目的" in header_text or "注意" in header_text):
                training_sheet = s
                break

    if mapping_sheet is None:
        for s in wb.sheetnames:
            rows = read_sheet_as_dicts(wb, s)
            if not rows:
                continue
            headers = rows[0].keys()
            header_text = " ".join(headers)
            if "チェック" in header_text and "トレーニング" in header_text:
                mapping_sheet = s
                break

    training_master = {}
    check_to_training = {}

    if training_sheet:
        rows = read_sheet_as_dicts(wb, training_sheet)

        for row in rows:
            tid = get_cell(row, ["トレーニングID", "TrainingID", "ID"])
            if not tid:
                continue

            title = get_cell(row, ["トレーニング名", "種目名", "メニュー名", "名前", "title"])
            purpose = get_cell(row, ["目的", "狙い", "効果"])
            notes = get_cell(row, ["注意点", "ポイント", "留意点"])
            count = get_cell(row, ["回数", "頻度", "セット", "時間"])
            gif_file = get_cell(row, ["GIF", "gif", "ファイル"])

            training_master[tid] = {
                "id": tid,
                "title": title or tid,
                "purpose": purpose,
                "notes": notes,
                "count": count,
                "gif_file": gif_file,
            }

    if mapping_sheet:
        rows = read_sheet_as_dicts(wb, mapping_sheet)

        for row in rows:
            cid_text = get_cell(row, ["チェックID", "CheckID", "評価ID", "項目ID", "C"])
            tid_text = get_cell(row, ["トレーニングID", "TrainingID", "T"])

            cids = split_ids(cid_text)
            tids = split_ids(tid_text)

            for cid in cids:
                if cid.startswith("C"):
                    check_to_training.setdefault(cid, [])
                    for tid in tids:
                        if tid.startswith("T") and tid not in check_to_training[cid]:
                            check_to_training[cid].append(tid)

    return training_master, check_to_training


def get_selected_training_ids(scores: dict) -> list[str]:
    training_master, check_to_training = load_training_data()

    selected_check_ids = [cid for cid, score in scores.items() if score >= 1]

    selected_training_ids = []
    for cid in selected_check_ids:
        for tid in check_to_training.get(cid, []):
            if tid not in selected_training_ids:
                selected_training_ids.append(tid)

    return selected_training_ids


def create_training_qr(url: str) -> Path:
    qr_path = OUTPUT_DIR / "selected_training_qr.png"
    img = qrcode.make(url)
    img.save(qr_path)
    return qr_path


def show_gif(check_id: str) -> None:
    gif_path = GIF_DIR / f"{check_id}.gif"
    if gif_path.exists():
        st.image(str(gif_path), use_container_width=True)
    else:
        st.info(f"{check_id} GIFなし：{gif_path.name}")


def show_ok_ng(check_id: str) -> None:
    ok_path = OK_NG_DIR / f"{check_id}ok.png"
    ng_path = OK_NG_DIR / f"{check_id}ng.png"

    col_ok, col_ng = st.columns(2)

    with col_ok:
        if ok_path.exists():
            st.image(str(ok_path), use_container_width=True)
        else:
            st.warning(f"{check_id} OK画像なし")

    with col_ng:
        if ng_path.exists():
            st.image(str(ng_path), use_container_width=True)
        else:
            st.warning(f"{check_id} NG画像なし")


def update_excel_scores(scores: dict) -> None:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE)
    ws = wb["評価入力"]

    headers = [cell.value for cell in ws[1]]
    check_id_col = headers.index("チェックID") + 1
    score_col = headers.index("スコア") + 1

    for row in range(2, ws.max_row + 1):
        check_id = ws.cell(row=row, column=check_id_col).value
        if check_id in scores:
            ws.cell(row=row, column=score_col).value = scores[check_id]

    wb.save(EXCEL_FILE)


def build_training_cards_html(selected_training_ids: list[str]) -> str:
    training_master, _ = load_training_data()

    if not selected_training_ids:
        return """
        <section class="training-card">
            <h3>自主トレーニングなし</h3>
            <p>全項目0点のため、追加で抽出された自主トレーニングはありません。</p>
        </section>
        """

    html = ""

    for tid in selected_training_ids:
        item = training_master.get(tid, {
            "id": tid,
            "title": tid,
            "purpose": "",
            "notes": "",
            "count": "",
            "gif_file": "",
        })

        gif_path = find_training_gif(tid, item.get("gif_file", ""))
        gif_src = image_to_base64_src(gif_path)

        if gif_src:
            gif_html = f'<img src="{gif_src}" alt="{escape(item["title"])}">'
        else:
            gif_html = f'<div class="training-no-gif">GIFなし<br>{escape(tid)}</div>'

        html += f"""
        <section class="training-card">
            <div class="training-id">{escape(tid)}</div>
            <h3>{escape(item["title"])}</h3>
            <div class="training-gif">
                {gif_html}
            </div>
            <div class="training-info">
                <h4>目的</h4>
                <p>{escape(item.get("purpose", "")).replace(chr(10), "<br>")}</p>
                <h4>注意点</h4>
                <p>{escape(item.get("notes", "")).replace(chr(10), "<br>")}</p>
                <h4>回数</h4>
                <p class="count">{escape(item.get("count", ""))}</p>
            </div>
        </section>
        """

    return html


def generate_selected_training_menu_html(scores: dict) -> Path:
    selected_training_ids = get_selected_training_ids(scores)
    training_cards_html = build_training_cards_html(selected_training_ids)

    html = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>自主トレーニングメニュー</title>
<style>
body {{
    margin: 0;
    padding: 14px;
    background: #f3f4f6;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    color: #111827;
}}
.container {{
    max-width: 760px;
    margin: 0 auto;
}}
.header {{
    background: linear-gradient(135deg, #111827, #374151);
    color: white;
    border-radius: 22px;
    padding: 22px 18px;
    margin-bottom: 16px;
}}
.header h1 {{
    margin: 0 0 8px;
    font-size: 25px;
}}
.header p {{
    margin: 0;
    opacity: 0.85;
}}
.training-card {{
    background: white;
    border-radius: 22px;
    padding: 18px;
    margin-bottom: 16px;
    box-shadow: 0 4px 14px rgba(0,0,0,0.08);
}}
.training-id {{
    display: inline-block;
    background: #111827;
    color: white;
    border-radius: 999px;
    padding: 5px 12px;
    font-size: 13px;
    font-weight: 900;
    margin-bottom: 8px;
}}
.training-card h3 {{
    margin: 0 0 12px;
    font-size: 23px;
}}
.training-gif {{
    text-align: center;
    margin: 12px 0 16px;
}}
.training-gif img {{
    max-width: 100%;
    width: 520px;
    border-radius: 14px;
}}
.training-no-gif {{
    background: #e5e7eb;
    border-radius: 14px;
    padding: 40px 12px;
    color: #6b7280;
    font-weight: 800;
}}
.training-info h4 {{
    margin: 16px 0 6px;
    font-size: 18px;
    border-left: 5px solid #111827;
    padding-left: 8px;
}}
.training-info p {{
    margin: 0;
    line-height: 1.7;
    font-size: 16px;
}}
.count {{
    font-weight: 900;
    background: #f9fafb;
    border-radius: 12px;
    padding: 12px;
}}
</style>
</head>
<body>
<div class="container">
    <section class="header">
        <h1>自主トレーニングメニュー</h1>
        <p>評価結果に基づいて抽出されたトレーニング一覧です。</p>
    </section>

    {training_cards_html}
</div>
</body>
</html>
"""

    html_path = OUTPUT_DIR / "selected_training_menu.html"
    html_path.write_text(html, encoding="utf-8")
    return html_path


def generate_mobile_html_report(scores: dict) -> tuple[Path, Path]:
    total_score = sum(scores.values())
    problem_checks = [c for c in CHECKS if scores.get(c["id"], 0) >= 1]
    selected_training_ids = get_selected_training_ids(scores)

    training_menu_path = generate_selected_training_menu_html(scores)
    training_menu_filename = training_menu_path.name

    # HTMLレポート内リンクは、同じoutputフォルダ内の selected_training_menu.html に飛ばす
    training_link_url = training_menu_filename

    qr_path = create_training_qr(training_link_url)
    qr_src = image_to_base64_src(qr_path)

    phase_names = ["ワインドアップ", "アーリーコッキング", "レイトコッキング", "アクセラレーション", "フォロースルー"]
    phase_problem_map = {phase: False for phase in phase_names}
    phase_check_map = {phase: [] for phase in phase_names}

    for check in problem_checks:
        phase_problem_map[check["phase"]] = True
        phase_check_map[check["phase"]].append(check["id"])

    display_phase_names = [phase for phase in phase_names if phase_problem_map[phase]]

    if total_score == 0:
        main_comment = "全体として大きなフォーム上の問題は確認されませんでした。現在のフォームを維持しながら、継続的なコンディショニングを行いましょう。"
    elif total_score <= 6:
        main_comment = "一部のフェーズで軽度のフォーム課題が確認されました。該当フェーズを中心に、動作の安定性と再現性を高めるトレーニングが推奨されます。"
    else:
        main_comment = "複数のフェーズでフォーム課題が確認されました。局所的な修正だけでなく、下肢・体幹・上肢の連動性を高める段階的なトレーニングが推奨されます。"

    phase_cards_html = ""

    if display_phase_names:
        for phase in display_phase_names:
            image_name = PHASE_IMAGE_MAP[phase]
            img_src = image_to_base64_src(PHASE_IMAGE_DIR / image_name)
            checks_text = " / ".join(phase_check_map[phase])

            image_html = f'<img src="{img_src}" alt="{escape(phase)}">' if img_src else f'<div class="no-image">画像なし<br>{escape(image_name)}</div>'

            phase_cards_html += f"""
            <section class="phase-card phase-problem">
                <div class="phase-image">
                    {image_html}
                    <div class="star">★</div>
                </div>
                <div class="phase-body">
                    <h3>{escape(phase)}</h3>
                    <p>該当チェック：{escape(checks_text)}</p>
                </div>
            </section>
            """
    else:
        phase_cards_html = """
        <section class="phase-card">
            <div class="phase-body">
                <h3>該当フェーズなし</h3>
                <p>フォーム上の大きな問題は確認されませんでした。</p>
            </div>
        </section>
        """

    problem_items_html = ""

    if problem_checks:
        for check in problem_checks:
            cid = check["id"]
            score = scores.get(cid, 0)
            criteria_text = escape(check["criteria"][score]).replace("\n", "<br>")

            problem_items_html += f"""
            <section class="problem-card">
                <div class="problem-header">
                    <span class="problem-id">{escape(cid)}</span>
                    <span class="problem-score">スコア {score}</span>
                </div>
                <h3>{escape(check["title"])}</h3>
                <p class="phase-label">{escape(check["phase"])}</p>
                <p class="criteria">{criteria_text}</p>
            </section>
            """
    else:
        problem_items_html = """
        <section class="problem-card good">
            <h3>チェック項目なし</h3>
            <p>全項目0点です。大きな問題はありません。</p>
        </section>
        """

    qr_html = f'<img class="qr-img" src="{qr_src}" alt="自主トレーニングQRコード">' if qr_src else '<div class="qr-missing">QR画像を作成できませんでした</div>'

    training_cards_html = build_training_cards_html(selected_training_ids)

    html = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>投球フォーム評価レポート</title>
<style>
body {{
    margin: 0;
    padding: 0;
    background: #f3f4f6;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    color: #111827;
}}
.container {{
    max-width: 720px;
    margin: 0 auto;
    padding: 14px;
    box-sizing: border-box;
}}
.header {{
    background: linear-gradient(135deg, #111827, #374151);
    color: white;
    border-radius: 22px;
    padding: 22px 18px;
    margin-bottom: 14px;
}}
.logo-row {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
}}
.logo {{
    width: 76px;
    height: 42px;
    border: 2px solid white;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 900;
    font-size: 20px;
}}
.header h1 {{
    margin: 18px 0 4px;
    font-size: 25px;
}}
.header p {{
    margin: 0;
    font-size: 13px;
    opacity: 0.85;
}}
.summary {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px;
    margin-bottom: 14px;
}}
.summary-card, .phase-card, .problem-card, .qr-card, .training-card {{
    background: white;
    border-radius: 18px;
    padding: 16px;
    box-shadow: 0 4px 14px rgba(0,0,0,0.08);
    margin-bottom: 14px;
}}
.summary-card .label {{
    font-size: 13px;
    color: #6b7280;
}}
.summary-card .value {{
    font-size: 30px;
    font-weight: 900;
    margin-top: 4px;
}}
.section-title {{
    margin: 24px 0 10px;
    font-size: 20px;
    border-left: 6px solid #111827;
    padding-left: 10px;
}}
.comment-card {{
    background: #fff7ed;
    border-left: 6px solid #f97316;
    border-radius: 18px;
    padding: 16px;
    line-height: 1.8;
    font-size: 15px;
}}
.phase-card {{
    padding: 0;
    overflow: hidden;
    border: 2px solid transparent;
}}
.phase-card.phase-problem {{
    border-color: #ef4444;
}}
.phase-image {{
    position: relative;
    background: #f9fafb;
    min-height: 150px;
    display: flex;
    align-items: center;
    justify-content: center;
}}
.phase-image img {{
    width: 100%;
    height: auto;
    display: block;
}}
.star {{
    position: absolute;
    top: 10px;
    right: 12px;
    color: #dc2626;
    font-size: 46px;
    font-weight: 900;
}}
.phase-body {{
    padding: 14px 16px;
}}
.phase-body h3 {{
    margin: 0 0 5px;
}}
.phase-body p, .phase-label {{
    margin: 0;
    color: #6b7280;
    font-weight: 700;
}}
.no-image, .training-no-gif {{
    width: 100%;
    min-height: 120px;
    display: flex;
    align-items: center;
    justify-content: center;
    text-align: center;
    color: #6b7280;
    font-weight: 700;
    background: #e5e7eb;
}}
.problem-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 8px;
}}
.problem-id, .training-id {{
    background: #111827;
    color: white;
    border-radius: 999px;
    padding: 5px 12px;
    font-size: 13px;
    font-weight: 800;
}}
.problem-score {{
    color: #b91c1c;
    font-weight: 900;
}}
.criteria {{
    line-height: 1.7;
    font-size: 15px;
}}
.qr-card {{
    text-align: center;
}}
.qr-img {{
    width: 190px;
    height: 190px;
    object-fit: contain;
    margin: 6px auto 12px;
    display: block;
}}
.training-link {{
    display: inline-block;
    margin-top: 14px;
    padding: 12px 18px;
    border-radius: 999px;
    background: #111827;
    color: white;
    text-decoration: none;
    font-weight: 900;
    font-size: 15px;
}}
.training-url {{
    margin-top: 10px;
    font-size: 12px;
    color: #2563eb;
    word-break: break-all;
}}
.training-card h3 {{
    margin: 8px 0 12px;
    font-size: 21px;
}}
.training-gif {{
    text-align: center;
    margin: 12px 0;
}}
.training-gif img {{
    max-width: 100%;
    width: 520px;
    border-radius: 14px;
}}
.training-info h4 {{
    margin: 14px 0 5px;
    font-size: 17px;
    border-left: 5px solid #111827;
    padding-left: 8px;
}}
.training-info p {{
    margin: 0;
    line-height: 1.7;
}}
.count {{
    font-weight: 900;
    background: #f9fafb;
    border-radius: 12px;
    padding: 10px;
}}
.footer {{
    text-align: center;
    color: #6b7280;
    font-size: 12px;
    margin: 24px 0 8px;
}}
</style>
</head>
<body>
<div class="container">

<header class="header">
    <div class="logo-row">
        <div class="logo">FGC</div>
        <div style="font-size:12px; opacity:0.8;">Pitching Training System</div>
    </div>
    <h1>投球フォーム評価レポート</h1>
    <p>Pitching Form Report & Training Program</p>
</header>

<section class="summary">
    <div class="summary-card">
        <div class="label">総合スコア</div>
        <div class="value">{total_score} / 22</div>
    </div>
    <div class="summary-card">
        <div class="label">チェック項目数</div>
        <div class="value">{len(problem_checks)}</div>
    </div>
</section>

<h2 class="section-title">総合コメント</h2>
<section class="comment-card">{escape(main_comment)}</section>

<h2 class="section-title">該当した投球フェーズ</h2>
{phase_cards_html}

<h2 class="section-title">チェックされたフォーム課題</h2>
{problem_items_html}

<h2 class="section-title">自主トレーニング</h2>
<section class="qr-card">
    {qr_html}
    <p>
        QRコード、または下のボタンから<br>
        評価結果に応じて抽出された自主トレーニング一覧を確認できます。
    </p>
    <a class="training-link" href="{escape(training_link_url)}" target="_blank">
        自主トレーニング一覧を開く
    </a>
    <div class="training-url">{escape(training_link_url)}</div>
    <p style="margin-top:10px; font-weight:800;">
        1点・2点のどちらでも、該当IDに紐づくトレーニングをすべて出力します。
    </p>
</section>

<h2 class="section-title">抽出された自主トレーニング一覧</h2>
{training_cards_html}

<div class="footer">Fukui General Clinic / FGC</div>

</div>
</body>
</html>
"""

    html_path = OUTPUT_DIR / "pitching_report_mobile.html"
    html_path.write_text(html, encoding="utf-8")
    return html_path, training_menu_path


st.set_page_config(
    page_title="投球フォーム評価",
    page_icon="⚾",
    layout="wide",
)


st.markdown(
    """
    <style>
    .stApp {
        background-color: #ffffff;
        color: #111827;
    }
    [data-testid="stHeader"] {
        background-color: #ffffff;
    }
    [data-testid="stSidebar"] {
        background-color: #ffffff;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 14px;
    }
    h1, h2, h3, h4, h5, h6, p, label, span {
        color: #111827;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


st.title("投球フォーム評価シート")
st.caption("GIF・OK/NG画像・評価基準を確認しながら、C01〜C11を0・1・2で評価します。")

st.markdown("---")

scores = {}

for check in CHECKS:
    cid = check["id"]

    with st.container(border=True):
        st.markdown(f"## {cid}　{check['title']}")
        st.caption(check["phase"])

        col_gif, col_img, col_score = st.columns([1.15, 1.55, 1.25])

        with col_gif:
            show_gif(cid)

        with col_img:
            show_ok_ng(cid)

        with col_score:
            st.markdown("### 評価基準")

            for score_value, comment in check["criteria"].items():
                st.markdown(f"**{score_value}点**")
                st.markdown(comment.replace("\n", "  \n"))

            selected_score = st.radio(
                "スコアを選択",
                options=[0, 1, 2],
                horizontal=True,
                key=f"score_{cid}",
            )

            scores[cid] = selected_score


st.markdown("---")
st.subheader("評価サマリー")

total_score = sum(scores.values())
problem_ids = [cid for cid, score in scores.items() if score >= 1]

col1, col2, col3 = st.columns(3)

with col1:
    st.metric("総合スコア", total_score)

with col2:
    st.metric("チェック項目数", len(problem_ids))

with col3:
    st.metric("最大スコア", 22)

if problem_ids:
    st.warning("チェックあり：" + " / ".join(problem_ids))
else:
    st.success("全項目0点です。大きな問題はありません。")


st.markdown("---")
st.subheader("出力設定")

output_type = st.radio(
    "出力形式を選択してください",
    ["スマホ用HTML", "PPTレポート", "両方"],
    horizontal=True,
)

if st.button("レポート作成", type="primary"):
    try:
        st.write("Excelにスコアを書き込み中...")
        update_excel_scores(scores)
        st.success("Excel更新完了")

        if output_type in ["スマホ用HTML", "両方"]:
            st.write("スマホ用HTMLレポートを作成中...")

            html_path, training_menu_path = generate_mobile_html_report(scores)

            st.success("スマホ用HTMLレポート完成！")
            st.info("HTMLレポート内に、評価結果で抽出された自主トレーニング一覧も表示されます。")

            with open(html_path, "rb") as f:
                st.download_button(
                    label="スマホ用HTMLレポートをダウンロード",
                    data=f,
                    file_name="pitching_report_mobile.html",
                    mime="text/html",
                )

            with open(training_menu_path, "rb") as f:
                st.download_button(
                    label="自主トレーニング一覧HTMLをダウンロード",
                    data=f,
                    file_name="selected_training_menu.html",
                    mime="text/html",
                )

            with st.expander("スマホ用HTMLプレビュー"):
                st.components.v1.html(
                    html_path.read_text(encoding="utf-8"),
                    height=1000,
                    scrolling=True,
                )

        if output_type in ["PPTレポート", "両方"]:
            st.write("PPTレポート生成中...")

            from report_generator_public import generate_report

            result = generate_report()

            st.success("PPTレポート完成！")

            pptx_path = result.get("pptx")
            pdf_path = result.get("pdf")

            if pptx_path and Path(pptx_path).exists():
                with open(pptx_path, "rb") as f:
                    st.download_button(
                        label="PPTXをダウンロード",
                        data=f,
                        file_name="pitching_report_auto.pptx",
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    )
            else:
                st.warning("PPTXが見つかりません。")

            if pdf_path and Path(pdf_path).exists():
                with open(pdf_path, "rb") as f:
                    st.download_button(
                        label="PDFをダウンロード",
                        data=f,
                        file_name="pitching_report_auto.pdf",
                        mime="application/pdf",
                    )
            else:
                st.info("PDFはクラウドでは作成されないため、PPTXをダウンロードしてください。")

            if result.get("url"):
                st.caption(f"QRリンク: {result['url']}")

    except Exception as e:
        st.error("レポート生成でエラーが発生しました")
        st.exception(e)